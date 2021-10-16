from os import error, path
import openpyxl as xl
from openpyxl import Workbook
import sys
import os




def generateBook(filename):
    # path = input('Enter Path(Enter "." for same location): ')
    # filename = input('Enter file name: ')
    book = Workbook()
    sheet = book.active
    sheetname = input('Enter active sheet name: ')
    sheet.title = sheetname
    book.save(f'{filename}.xlsx')
    book.close()

def showBook(bookname):
    try:   
        sheet = bookname.active
        row = sheet.max_row
        col = sheet.max_column
        for i in range(1,row+1):
            for j in range(1,col+1):
                print(f'{i} : {j} => {sheet.cell(row=i, column=j).value}')
    
    except(error):
        print(error)


def addtobook(bookname, filename):
    sheet = bookname.active
    if sheet.max_column == 0:
        col = input('Enter column length: ')
        row = sheet.max_row
    else:
        row = sheet.max_row
        col = sheet.max_column
    for i in range(1, col+1):
        data = input(f'Enter value at row: {row+1} and column : {i}: ')
        sheet.cell(row = row+1, column=col).value = data
    bookname.save(filename)
    bookname.close()

def modifyrecord(bookname, filename):
    row = input('Enter Row: ')
    col = input('Enter Column: ')
    data = input("Enter data: ")
    sheet = bookname.active
    sheet.cell(row=row, column=col).value = ''
    sheet.cell(row=row, column=col).value = data
    bookname.save(filename)

def deleterecord(bookname, filename):
    row = int(input('Enter Row: '))
    sheet = bookname.active
    check =  input('Are you want to delete(Y|N): ')
    check = check.upper()
    if check == 'Y':
        sheet.delete_rows(row,1)
        bookname.save(filename)
    else:
        pass
    
def deletecolumn(bookname, filename):
    col = int(input('Enter Column: '))
    sheet = bookname.active
    check = input('Are you want to delete(Y|N): ')
    check = check.upper()
    if check == 'Y':
        sheet.delete_columns(col , 1)
        bookname.save(filename)
    else:
        pass

def insertcolumn(bookname, filename):
    sheet = bookname.active
    columnlength = sheet.max_column
    sheet.insert_cols(columnlength+1, 1)
    columnname = input('Enter Column Name: ')
    sheet.cell(1,columnlength+1).value = columnname
    bookname.save(filename)
    
def setdefaultbook(filename):
    path = os.path(filename)
    # filename = input('Enter file name: ')
    # path = input('Enter file path: ')
    try:
        defaultfile = open('defaultfile.txt', 'w')
        defaultfile.seek(0)
        defaultfile.truncate()
        defaultfile.write(f'{path}')
    except(error):
        print(error)

def opendefaultbook():
    try:
        file = open('defaultfile.txt', 'r')
        path = file.read()
        filename = os.path.basename(path)
        path = os.path.dirname(path)
        book = xl.load_workbook(f'{filename}')
        print(f'{filename} is open')
        return book
    except(error):
        print(error)


filename = ''
book = ''
path = ''

def main(book, filename):
    global path
    cmd = sys.argv[1]
    if cmd.lower( )== 'select':
        path = sys.argv[2]
        filename = os.path.basename(path)
        path = os.path.dirname(path)
        print(f'{filename} is selected!')
        try:
            book=xl.load_workbook(f'{filename}.xlsx')
        except(error):
            print(error)
    elif cmd.lower() == '-o':
        book = opendefaultbook()
    check = True
    while check:
        command = input('>>>')
        check = commands(book,filename,command, path)
        if not check:
            book.close()
            break

def help():
    print('''
        xcel select bookname => open book
        xcel -g => generate new book
        xcel -s => show data in active book sheet
        xcel -a => add data to active book sheet
        xcel -m => mpdify data
        xcel -dr => delete a record/row 
        xcel -dc => delete a column
        xcel -i => insert new column
        xcel -d => set book as default
        xcel -o => open default book
        exit => exit from selected book
        ''')


def commands(book, filename, cmd, path):
    print(cmd.lower())
    if cmd.lower() == 'xcel help':
        help()
    elif cmd.lower() == 'xcel -g':
        generateBook(sys.argv[2])

    elif cmd.lower() == 'xcel -s':
        showBook(bookname=book)

    elif cmd.lower() == 'xcel -a':
        addtobook(bookname=book, filename=filename)

    elif cmd.lower() == 'xcel -m':
        modifyrecord(bookname=book, filename=filename)

    elif cmd.lower == 'xcel -dr':
        deleterecord(bookname=book, filename=filename)

    elif cmd.lower() == 'xcel -dc':
        deletecolumn(bookname=book, filename=filename)

    elif cmd.lower() == 'xcel -i':
        insertcolumn(bookname=book, filename=filename)

    elif cmd.lower() == 'xcel -d':
        setdefaultbook(filename=filename)
    elif cmd.lower() == 'exit':
        return False

    else:
        print('no command found')
        return False
    return True

if __name__ == '__main__':
    print('Type "xcel help" for help')
    cmd = sys.argv[1]
    if cmd.lower() == 'help':
        help()
    else:
        main(book,filename)
        

